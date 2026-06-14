import os
import datetime
import logging
import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session
from database import SessionLocal, TranslationMemory, Glossary, TranslationJob
from services.gemini_service import GeminiService

logger = logging.getLogger("locflow.excel")

class ExcelService:
    def __init__(self):
        self.gemini_service = GeminiService()

    def import_dictionary_file(self, filepath: str) -> int:
        """
        Imports Chinese-Vietnamese translation pairs from an Excel file into Translation Memory.
        Ignores rows with duplicate source text.
        """
        db: Session = SessionLocal()
        inserted_count = 0
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active  # Uses the active sheet
            
            for row in range(2, ws.max_row + 1):  # Skip Header Row 1
                src = ws.cell(row=row, column=1).value
                tgt = ws.cell(row=row, column=2).value
                
                if src is not None and tgt is not None:
                    src_str = str(src).strip()
                    tgt_str = str(tgt).strip()
                    
                    if not src_str or not tgt_str:
                        continue
                    
                    # Check if source text already exists in Translation Memory or Glossary
                    exists_in_tm = db.query(TranslationMemory).filter(
                        TranslationMemory.source_text == src_str
                    ).first()
                    exists_in_glossary = db.query(Glossary).filter(
                        Glossary.source_term == src_str
                    ).first()
                    
                    if not exists_in_tm and not exists_in_glossary:
                        tm_entry = TranslationMemory(
                            source_text=src_str,
                            translated_text=tgt_str
                        )
                        db.add(tm_entry)
                        inserted_count += 1
                        
            db.commit()
            logger.info(f"Imported {inserted_count} translation memory entries.")
            return inserted_count
        except Exception as e:
            db.rollback()
            logger.error(f"Error importing dictionary: {e}")
            raise e
        finally:
            db.close()

    def import_glossary_file(self, filepath: str) -> int:
        """
        Imports terminology entries from a 2-column Excel file into the Glossary database.
        Ignores duplicates.
        """
        db: Session = SessionLocal()
        inserted_count = 0
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):  # Skip Header Row 1
                src = ws.cell(row=row, column=1).value
                tgt = ws.cell(row=row, column=2).value
                
                if src is not None and tgt is not None:
                    src_str = str(src).strip()
                    tgt_str = str(tgt).strip()
                    
                    if not src_str or not tgt_str:
                        continue
                    
                    exists_in_glossary = db.query(Glossary).filter(
                        Glossary.source_term == src_str
                    ).first()
                    exists_in_tm = db.query(TranslationMemory).filter(
                        TranslationMemory.source_text == src_str
                    ).first()
                    
                    if not exists_in_glossary and not exists_in_tm:
                        glossary_entry = Glossary(
                            source_term=src_str,
                            target_term=tgt_str
                        )
                        db.add(glossary_entry)
                        inserted_count += 1
            
            db.commit()
            logger.info(f"Imported {inserted_count} glossary terms.")
            return inserted_count
        except Exception as e:
            db.rollback()
            logger.error(f"Error importing glossary: {e}")
            raise e
        finally:
            db.close()

    def process_translation_job(self, job_id: str, input_filepath: str, output_filepath: str, save_to_tm: bool = True):
        """
        Processes translation on the uploaded Excel in a background thread.
        Reads row by row, looks up TM, batch-translates via Gemini, format-styles AI cells,
        and saves intermediate states.
        """
        db: Session = SessionLocal()
        try:
            # 1. Update job status to PROCESSING
            job = db.query(TranslationJob).filter(TranslationJob.id == job_id).first()
            if not job:
                logger.error(f"Job {job_id} not found in database.")
                return
            
            job.status = "PROCESSING"
            db.commit()

            # 2. Open Workbook (preserving styles and formulas)
            wb = openpyxl.load_workbook(input_filepath)
            
            # 3. Identify all rows needing translation across all sheets
            rows_to_translate = []  # List of dicts: {"sheet_name": str, "row_idx": int, "source_text": str}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r in range(2, ws.max_row + 1):  # Skip header (Row 1)
                    src_val = ws.cell(row=r, column=1).value
                    tgt_val = ws.cell(row=r, column=2).value
                    
                    if src_val is not None:
                        src_str = str(src_val).strip()
                        # Process if Chinese is present and Vietnamese is blank
                        if src_str and (tgt_val is None or str(tgt_val).strip() == ""):
                            rows_to_translate.append({
                                "sheet_name": sheet_name,
                                "row_idx": r,
                                "source_text": src_str
                            })

            total_rows = len(rows_to_translate)
            job.total_rows = total_rows
            db.commit()
            
            if total_rows == 0:
                # Nothing to translate
                job.status = "COMPLETED"
                job.completed_at = datetime.datetime.utcnow()
                db.commit()
                wb.save(output_filepath)
                return

            # Load the entire glossary to scan efficiently in memory
            all_glossary = db.query(Glossary).all()
            
            batch_size = 30
            processed_count = 0
            tm_hits_count = 0
            ai_trans_count = 0
            
            # Temp store for items waiting for AI translation
            ai_batch_items = []  # list of dicts

            def flush_ai_batch():
                nonlocal ai_trans_count
                if not ai_batch_items:
                    return
                
                texts_to_translate = [item["source_text"] for item in ai_batch_items]
                
                # Filter glossary terms that actually appear in this batch of text
                combined_texts = " ".join(texts_to_translate)
                active_glossary = []
                for entry in all_glossary:
                    if entry.source_term in combined_texts:
                        active_glossary.append((entry.source_term, entry.target_term))
                
                try:
                    translated_list = self.gemini_service.translate_batch(texts_to_translate, active_glossary)
                except Exception as api_err:
                    logger.error(f"Gemini batch translation failed: {api_err}")
                    raise api_err
                
                # Write results back to cells and save to TM
                for idx, translated_text in enumerate(translated_list):
                    item = ai_batch_items[idx]
                    sheet = wb[item["sheet_name"]]
                    cell = sheet.cell(row=item["row_idx"], column=2)
                    
                    cell.value = translated_text
                    cell.font = Font(bold=True)  # Format AI translated cells as BOLD
                    
                    # Store new translation in TM database if not existing and save_to_tm is enabled
                    if save_to_tm:
                        tm_exists = db.query(TranslationMemory).filter(
                            TranslationMemory.source_text == item["source_text"]
                        ).first()
                        
                        if not tm_exists:
                            new_tm = TranslationMemory(
                                source_text=item["source_text"],
                                translated_text=translated_text
                            )
                            db.add(new_tm)
                    
                    ai_trans_count += 1

                # Clean up batch list
                ai_batch_items.clear()
                db.commit()

            # Iterate through rows needing translation
            for row_info in rows_to_translate:
                src_text = row_info["source_text"]
                
                # Check Translation Memory first
                tm_record = db.query(TranslationMemory).filter(
                    TranslationMemory.source_text == src_text
                ).first()
                
                # Check Glossary next for exact match
                gloss_record = None
                if not tm_record:
                    gloss_record = db.query(Glossary).filter(
                        Glossary.source_term == src_text
                    ).first()
                
                if tm_record:
                    # TM Hit
                    sheet = wb[row_info["sheet_name"]]
                    cell = sheet.cell(row=row_info["row_idx"], column=2)
                    cell.value = tm_record.translated_text
                    # Note: We do NOT bold TM hits (we preserve normal font)
                    
                    tm_hits_count += 1
                    processed_count += 1
                elif gloss_record:
                    # Glossary Direct Hit (Reuse immediately without calling AI)
                    sheet = wb[row_info["sheet_name"]]
                    cell = sheet.cell(row=row_info["row_idx"], column=2)
                    cell.value = gloss_record.target_term
                    
                    tm_hits_count += 1
                    processed_count += 1
                else:
                    # AI Translate Needed
                    ai_batch_items.append(row_info)
                    processed_count += 1
                    
                    # If batch is full, translate
                    if len(ai_batch_items) >= batch_size:
                        flush_ai_batch()
                
                # Periodically update progress in the database (every 10 rows)
                if processed_count % 10 == 0 or processed_count == total_rows:
                    job.processed_rows = processed_count
                    job.tm_hits = tm_hits_count
                    job.ai_translations = ai_trans_count
                    db.commit()

            # Flush any remaining items in the AI queue
            if ai_batch_items:
                flush_ai_batch()
                # Final count updates
                job.processed_rows = processed_count
                job.tm_hits = tm_hits_count
                job.ai_translations = ai_trans_count
                db.commit()

            # Create output directories if needed
            out_dir = os.path.dirname(output_filepath)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            
            # Save the updated Excel workbook
            wb.save(output_filepath)
            
            # Mark job complete
            job.status = "COMPLETED"
            job.completed_at = datetime.datetime.utcnow()
            db.commit()
            logger.info(f"Job {job_id} completed successfully. Total processed: {processed_count}")
            
        except Exception as e:
            db.rollback()
            logger.error(f"Error processing translation job {job_id}: {e}")
            job = db.query(TranslationJob).filter(TranslationJob.id == job_id).first()
            if job:
                job.status = "FAILED"
                job.error_message = str(e)
                job.completed_at = datetime.datetime.utcnow()
                db.commit()
        finally:
            db.close()

    def export_glossary_file(self, filepath: str):
        """
        Exports all Glossary terms to a 2-column Excel file.
        """
        db: Session = SessionLocal()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Glossary"
            ws.append(["Chinese", "Vietnamese"]) # Headers
            
            # Query all terms
            terms = db.query(Glossary).all()
            for term in terms:
                ws.append([term.source_term, term.target_term])
                
            wb.save(filepath)
        except Exception as e:
            logger.error(f"Error exporting glossary: {e}")
            raise e
        finally:
            db.close()

    def export_memory_file(self, filepath: str):
        """
        Exports all Translation Memory entries to a 2-column Excel file.
        """
        db: Session = SessionLocal()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TranslationMemory"
            ws.append(["Chinese", "Vietnamese"]) # Headers
            
            # Query all items
            items = db.query(TranslationMemory).all()
            for item in items:
                ws.append([item.source_text, item.translated_text])
                
            wb.save(filepath)
        except Exception as e:
            logger.error(f"Error exporting memory: {e}")
            raise e
        finally:
            db.close()

