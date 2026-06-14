import os
import sys

# Reconfigure stdout to support printing Vietnamese unicode characters on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Force test database URL before importing database module
os.environ["DATABASE_URL"] = "sqlite:///./test_locflow.db"

import shutil
import uuid
import openpyxl
from openpyxl.styles import Font
from database import init_db, SessionLocal, TranslationMemory, Glossary, TranslationJob
from services.excel_service import ExcelService
from services.gemini_service import GeminiService

class MockGeminiService:
    """Mock Gemini client to run validation tests offline without an API key."""
    def translate_batch(self, texts, glossary_pairs=None):
        translations = []
        glossary_dict = dict(glossary_pairs) if glossary_pairs else {}
        
        for text in texts:
            # Check glossary mapping first
            if text in glossary_dict:
                translations.append(glossary_dict[text])
            elif text == "背包":
                translations.append("Túi đồ")
            elif text == "登录":
                translations.append("Đăng nhập")
            elif text == "游戏开始":
                translations.append("Bắt đầu trò chơi")
            else:
                translations.append(f"[Mocked translation of {text}]")
        return translations

def run_test():
    print("--- Starting Backend Pipeline Test ---")
    
    # 1. Setup temporary database environment
    os.environ["DATABASE_URL"] = "sqlite:///./test_locflow.db"
    if os.path.exists("test_locflow.db"):
        os.remove("test_locflow.db")
        
    init_db()
    db = SessionLocal()
    
    # 2. Seed Glossary and Translation Memory
    print("[Test] Seeding Glossary and Translation Memory...")
    # Add a glossary term (Chinese "登录" -> Vietnamese "Đăng nhập")
    glossary_item = Glossary(source_term="登录", target_term="Đăng nhập")
    db.add(glossary_item)
    
    # Add a Translation Memory hit (Chinese "游戏开始" -> Vietnamese "Bắt đầu trò chơi")
    tm_item = TranslationMemory(source_text="游戏开始", translated_text="Bắt đầu trò chơi")
    db.add(tm_item)
    db.commit()
    
    # 3. Create mock input Excel file
    print("[Test] Creating Mock Excel File...")
    input_file = "test_input.xlsx"
    output_file = "test_output.xlsx"
    
    wb = openpyxl.Workbook()
    # Sheet 1: UI
    ws1 = wb.active
    ws1.title = "UI_Strings"
    ws1.append(["Chinese", "Vietnamese"]) # Headers
    ws1.append(["游戏开始", ""])  # Row 2 (Should use Translation Memory hit)
    ws1.append(["背包", ""])      # Row 3 (Should translate via Mock AI)
    ws1.append(["登录", ""])      # Row 4 (Should use Glossary term via Mock AI)
    ws1.append(["退出", "Thoát"])  # Row 5 (Should skip - already translated)
    
    # Sheet 2: Quests
    ws2 = wb.create_sheet(title="Quest_Text")
    ws2.append(["Chinese", "Vietnamese"]) # Headers
    ws2.append(["任务1", ""])      # Row 2 (Should translate via Mock AI)
    
    wb.save(input_file)
    
    # 4. Initialize service and inject Mock Gemini Service
    excel_service = ExcelService()
    excel_service.gemini_service = MockGeminiService()
    
    # 5. Create translation job
    job_id = str(uuid.uuid4())
    job = TranslationJob(
        id=job_id,
        filename="test_input.xlsx",
        status="PENDING"
    )
    db.add(job)
    db.commit()
    
    # 6. Run process job
    print("[Test] Running translation job background worker...")
    excel_service.process_translation_job(job_id, input_file, output_file)
    
    # Refresh database object
    db.refresh(job)
    print(f"[Test] Job status: {job.status}")
    print(f"[Test] Total rows target: {job.total_rows}")
    print(f"[Test] Processed rows: {job.processed_rows}")
    print(f"[Test] Translation Memory hits: {job.tm_hits}")
    print(f"[Test] AI translations: {job.ai_translations}")
    
    # Assert database states
    assert job.status == "COMPLETED", "Job failed to complete"
    assert job.total_rows == 4, f"Expected 4 rows needing translation, got {job.total_rows}"
    assert job.tm_hits == 2, f"Expected 2 TM/Glossary hits, got {job.tm_hits}"
    assert job.ai_translations == 2, f"Expected 2 AI translations, got {job.ai_translations}"
    
    # 7. Validate generated Excel file properties
    print("[Test] Verifying output Excel formatting...")
    wb_out = openpyxl.load_workbook(output_file)
    
    # Validate sheet UI_Strings
    ws1_out = wb_out["UI_Strings"]
    
    # Row 2: "游戏开始" -> "Bắt đầu trò chơi" (TM hit, should NOT be bold)
    r2_val = ws1_out.cell(row=2, column=2).value
    r2_bold = ws1_out.cell(row=2, column=2).font.bold
    print(f"  Row 2 (TM Hit): {r2_val} (Bold: {r2_bold})")
    assert r2_val == "Bắt đầu trò chơi", "TM hit value incorrect"
    assert r2_bold is not True, "TM hits should not be bolded"
    
    # Row 3: "背包" -> "Túi đồ" (AI translated, should be bold)
    r3_val = ws1_out.cell(row=3, column=2).value
    r3_bold = ws1_out.cell(row=3, column=2).font.bold
    print(f"  Row 3 (AI Translate): {r3_val} (Bold: {r3_bold})")
    assert r3_val == "Túi đồ", "AI translation value incorrect"
    assert r3_bold is True, "AI translations must be bolded"

    # Row 4: "登录" -> "Đăng nhập" (Glossary direct hit, should NOT be bold)
    r4_val = ws1_out.cell(row=4, column=2).value
    r4_bold = ws1_out.cell(row=4, column=2).font.bold
    print(f"  Row 4 (Glossary Hit): {r4_val} (Bold: {r4_bold})")
    assert r4_val == "Đăng nhập", "Glossary match value incorrect"
    assert r4_bold is not True, "Glossary matches should not be bolded"
    
    # Row 5: "退出" -> "Thoát" (Pre-existing, should be unchanged and not bold)
    r5_val = ws1_out.cell(row=5, column=2).value
    r5_bold = ws1_out.cell(row=5, column=2).font.bold
    print(f"  Row 5 (Pre-existing): {r5_val} (Bold: {r5_bold})")
    assert r5_val == "Thoát", "Pre-existing value changed"
    assert r5_bold is not True, "Pre-existing cells should not be bolded"
    
    # Validate sheet Quest_Text
    ws2_out = wb_out["Quest_Text"]
    # Row 2: "任务1" -> "[Mocked translation of 任务1]" (AI, should be bold)
    r2_q_val = ws2_out.cell(row=2, column=2).value
    r2_q_bold = ws2_out.cell(row=2, column=2).font.bold
    print(f"  Quest Row 2 (AI): {r2_q_val} (Bold: {r2_q_bold})")
    assert r2_q_val == "[Mocked translation of 任务1]", "Quest translation value incorrect"
    assert r2_q_bold is True, "Quest AI translation must be bolded"

    # 8. Clean up test files
    db.close()
    
    from database import engine
    engine.dispose()  # Releases file locks on Windows
    
    if os.path.exists("test_locflow.db"):
        os.remove("test_locflow.db")
    if os.path.exists(input_file):
        os.remove(input_file)
    if os.path.exists(output_file):
        os.remove(output_file)
        
    print("\n--- Pipeline validation passed successfully! ---")

if __name__ == "__main__":
    run_test()
